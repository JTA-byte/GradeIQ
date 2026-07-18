"""
Exports every row in the Supabase `cards` table to an Excel file.

Run:
  cd supabase
  python export_cards.py

Requires: openpyxl, supabase, python-dotenv (same packages already
pinned in python-services/requirements.txt):
  pip install -r ../python-services/requirements.txt

Reads NEXT_PUBLIC_SUPABASE_URL (falling back to SUPABASE_URL) and
SUPABASE_SERVICE_ROLE_KEY from ../.env.local -- same convention as
seed_cards.py in this directory.
"""
from __future__ import annotations

import os
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from supabase import Client, create_client

# .env.local lives one directory up from this script (gradeiq/.env.local).
load_dotenv(Path(__file__).resolve().parent.parent / ".env.local")

# PostgREST (Supabase's API layer) caps any unpaginated request at 1000
# rows by default -- paginate past it rather than silently truncating
# the export for a `cards` table bigger than that (confirmed necessary:
# this table has 5000+ rows after supabase/seed_cards.py was run).
PAGE_SIZE = 1000
OUTPUT_FILE = Path(__file__).resolve().parent / "cards_export.xlsx"
HEADERS = ["Name", "Set Name", "Card Number", "Rarity"]


def get_supabase_client() -> Client:
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError(
            "NEXT_PUBLIC_SUPABASE_URL (or SUPABASE_URL) and SUPABASE_SERVICE_ROLE_KEY "
            "must be set in gradeiq/.env.local"
        )
    return create_client(url, key)


def fetch_all_cards(client: Client) -> list[dict]:
    all_cards: list[dict] = []
    offset = 0

    while True:
        response = (
            client.table("cards")
            .select("name, set_name, card_number, rarity")
            .order("name")
            .range(offset, offset + PAGE_SIZE - 1)
            .execute()
        )
        page = response.data
        all_cards.extend(page)

        if len(page) < PAGE_SIZE:
            break
        offset += PAGE_SIZE

    return all_cards


def export_to_excel(cards: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cards"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for card in cards:
        ws.append(
            [card.get("name"), card.get("set_name"), card.get("card_number"), card.get("rarity")]
        )

    # Auto-size each column to its longest value (+ a little padding),
    # capped so one unusually long value doesn't blow out the width.
    for col_idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(output_path)


def main() -> None:
    client = get_supabase_client()
    print("Fetching cards from Supabase...")
    cards = fetch_all_cards(client)
    print(f"Fetched {len(cards)} cards.")

    export_to_excel(cards, OUTPUT_FILE)
    print(f"Exported {len(cards)} cards to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
